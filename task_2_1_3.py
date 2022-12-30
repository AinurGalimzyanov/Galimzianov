import csv
import math
from _datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
import datetime as dt
import dateparser
import cProfile

def profile(func):
    """Decorator for run function profile"""
    def wrapper(*args, **kwargs):
        profile_filename = func.__name__ + '.prof'
        profiler = cProfile.Profile()
        result = profiler.runcall(func, *args, **kwargs)
        profiler.dump_stats(profile_filename)
        return result
    return wrapper

currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
    }


class Salary:
    """
        Класс для представления зарплаты
        :param salary_in_rur: оклад в рублях
        :type salary_in_rur: float
    """
    def __init__(self, dictionary):
        """
        Инициализирует объект класса Salary
        :param dictionary: словарь с данными о окладе
        :type dictionary: dict
        """
        self.salary_in_rur = currency_to_rub[dictionary["salary_currency"]] * \
                        (math.floor(float(dictionary["salary_to"])) + math.floor(float(dictionary["salary_from"])))/2


class Vacancy:
    """
    Класс для представления вакансий
    :param dic: словарь с данными о вакансии
    :type dic: dict
    :param salary: оклад вакансий
    :type salary: Salary
    :param is_needed: соответствет ли вокансия введенной профессии
    :type is_needed: bool
    """
    def __init__(self, dictionary: dict):
        """
        Инициализирует объект класса Vacancy
        :param dictionary: словарь с данными о вакансияъ
        :type dictionary: dict
        """
        self.dic = dictionary
        self.salary = Salary(dictionary)
        self.dic["year"] = ParserDT(dictionary["published_at"]).get_year_one()
        self.is_needed = dictionary["is_needed"]


class ParserDT:
    """
    Класс для парсинга строки даты и времени
    :param datetime: полная строка даты и времени
    :type datetime: str
    """
    def __init__(self, str_datetime: str):
        """
        Инициализирует объект класса DateTimeParser
        :param str_datetime: полная строка даты и времени
        :type str_datetime: str
        """
        self.datetime = str_datetime

    def get_year_one(self):
        """
        Функция для получения года с помощью класса datetime
        :return: Год
        :rtype: int
        """
        format = '%Y-%m-%dT%H:%M:%S%z'
        return dt.datetime.strptime(self.datetime, format).year

    def get_year_two(self):
        """
        Функция для получения года с помощью метода split
        :return: Год
        :rtype: str
        """
        return self.datetime.split('-')[0]

    def get_year_three(self):
        """
        Функция для получения года с помощью индексации строк
        :return: Год
        :rtype: str
        """
        return self.datetime[:4]

    def get_year_four(self):
        """
        Функция для получения года с помощью класса dateparser
        :return: Год
        :rtype: int
        """
        return str(dateparser.parse(self.datetime).year)

class InputConect:
    """
    Класс представляющий входные данные
    :param file: имя файла
    :type file: str
    :param prof: название введеной профессии
    :type prof: str
    """
    def __init__(self):
        """
        Инициализирует объект класса InputConect
        """
        self.file = input("Введите название файла: ")
        self.prof = input("Введите название профессии: ")
        with open(self.file, "r", encoding='utf-8-sig', newline='') as test:
            unpacker = iter(csv.reader(test))
            if next(unpacker, "none") == "none":
                print("Пустой файл")
                exit(0)
            if next(unpacker, "none") == "none":
                print("Нет данных")
                exit(0)

class DataSet:
    """
    Класс представляющий собранные данные о ваканиях
    :param input_values: входные данные
    :type input_values: InputConect
    :param years: годы
    :type years: list
    :param wages_year: словарь заработныч плат за все года
    :type wages_year: dict
    :param price_area: словарь зарплат по городам
    :type price_area: dict
    :param project_list: лист словарей для передачи в класс Report
    :type project_list: list
    :param count_year: словарь количества вакансий за все года
    :type count_year: dict
    :param count_area: словарь количества вакансий по городам
    :type count_area:dict
    """
    def __init__(self):
        """
        Инициализирует объект класса DataSet
        """
        self.input_values = InputConect()
        self.csv_reader()
        self.csv_filter()
        self.years = []

        for i in self.filt_vac:
            self.years.append(i.dic["year"])

        self.years.sort()
        conut_vacs = len(self.filt_vac)
        self.wages_year, self.count_year = self.get_key_and_count(self.filt_vac, "year", False)
        self.year_required_wages, self.year_required_count = self.get_key_and_count(self.required_vac, "year", False)
        self.wages_area, self.count_area = self.get_key_and_count(self.filt_vac, "area_name", True)
        self.wages_area = dict(list(sorted(self.wages_area.items(), key=lambda x: x[1], reverse=True))[:10])
        self.price_area = {x: round(y / conut_vacs, 4) for x, y in self.count_area.items()}
        self.price_area = dict(list(sorted(self.price_area.items(), key=lambda x: x[1], reverse=True))[:10])
        self.project_list = self.creat_list()
        self.write_data()


    def csv_reader(self):
        """
        Читает csv файл
        :return: заполнил массив данных о вакансиях
        """
        with open(self.input_values.file, "r", encoding='utf-8-sig', newline='') as test:
            csv_file = csv.reader(test)
            self.first = next(csv_file)
            self.lines = []
            lenght = 0
            for row in csv_file:
                if lenght < len(row):
                    lenght = len(row)
                if not "" in row and lenght == len(row):
                    self.lines.append(row)

    def csv_filter(self):
        """
        Отбирает вакансии соответствующие введенной профессии
        :return: отфильтрованные массив вакансий
        """
        self.filt_vac = []
        pr = cProfile.Profile()
        pr.enable()
        for i in self.lines:
            new_dict = dict(zip(self.first, i))
            new_dict["is_needed"] = new_dict["name"].find(self.input_values.prof) > -1
            self.filt_vac.append(Vacancy(new_dict))
        self.required_vac = list(filter(lambda x: x.is_needed, self.filt_vac))
        pr.disable()
        pr.print_stats()

    def try_add(self, dic: dict, key, x):
        """
        Обновляет словарь данных статистики
        :param dic: словарь
        :type dic:dict
        :param key: ключ
        :type key: str
        :param x: значение
        :type x: float
        :return: обновленные словарь
        :rtype: dict
        """
        try: dic[key] += x
        except: dic[key] = x
        return dic

    def upgrad_keys(self, count_key):
        """
        Воозвращает пустой словарь с входными ключами
        :param count_key: входные ключи
        :type count_key: dict
        :return: Воозвращает пустой словарь с входными ключами
        :rtype: dict
        """
        for i in self.years:
            if i in count_key.keys():
                continue
            count_key[i] = 0
        return count_key

    def get_key_and_count(self, list_vacs: list, str_key: str, is_area: bool):
        """
        Возвращает словари с данными о зарплатах и количествах вакансий
        :param list_vacs: массив вакансий
        :type list_vacs: list
        :param str_key: ключ сбора статистики
        :type str_key:str
        :param is_area: статистика для городов или нет
        :type is_area:bool
        :return: словари с данными о вакансиях
        :rtype: tuple
        """
        sum_key, count_key = {}, {}
        for x in list_vacs:
            sum_key = self.try_add(sum_key, x.dic[str_key], x.salary.salary_in_rur)
            count_key = self.try_add(count_key, x.dic[str_key], 1)
        if not is_area:
            sum_key = self.upgrad_keys(sum_key)
            count_key = self.upgrad_keys(count_key)
        else:
            count_key = dict(filter(lambda y: y[1] / len(list_vacs) > 0.01, count_key.items()))
        salary_key = {}
        for key, x in count_key.items():
            salary_key[key] = 0 if x == 0 else math.floor(sum_key[key] / x)
        key_to_middle_salary = salary_key
        return key_to_middle_salary, count_key
    def write_data(self):
        """
        Печатает обработанные данные
        :return: напечатанные данные
        """
        print("Динамика уровня зарплат по годам:", self.wages_year)
        print("Динамика количества вакансий по годам:", self.count_year)
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.year_required_wages)
        print("Динамика количества вакансий по годам для выбранной профессии:", self.year_required_count)
        print("Уровень зарплат по городам (в порядке убывания):", self.wages_area)
        print("Доля вакансий по городам (в порядке убывания):", self.price_area)

    @profile
    def creat_list(self):
        """Создает лист словарей для передачи в класс Report

        :return: лист словарей
        :rtype: list
        """
        return [self.wages_year, self.year_required_wages, self.count_year,
                self.year_required_count, self.wages_area, self.price_area]

class Report:
    """
    Класс представляющий отчет
    :param dataSet_list: лист словарей с данными
    :type dataSet_list: list
    :param sheet_year: exl лист с таблицей статистики по годам
    :type sheet_year: worksheet
    :param sheet_city: exl лист с таблицами статистики по городам
    :type sheet_city: worksheet
    """
    def __init__(self):
        """
        Инициализирует объект класса Report
        """
        self.dataSet_list = DataSet().project_list
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.sheet_year = self.workbook.create_sheet("Статистика по годам")
        self.sheet_city = self.workbook.create_sheet("Статистика по городам")

    def generate_excel(self):
        """
        Создает таблицу статистики в xlsx файл

        :return: возвращиет xlsx файл
        """
        def creator_column_name(array, sheet):
            """
            Создает в определенный sheet имена column

            :param array: массив столбцов таблицы
            :type array: list

            :param sheet: exl лист книги
            :type array: worksheet

            :return: обозвал ячейки шапки таблицы
            """
            for index, column_name in enumerate(array):
                sheet.cell(row=1, column=index + 1, value=column_name).font = Font(bold=True)

        def setting_column_width_and_border_style(sheet_year) -> None:
            """
             В exl листе создает ширину ячеек и их дизайн

            :param sheet_year: exl лист с таблицей статистики по годам
            :type sheet_year: worksheet

            :return:
            """
            for column_cells in sheet_year.columns:
                length = 0
                for cell in column_cells:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    if cell.value is not None:
                        length = max(len(str(cell.value)), length)
                sheet_year.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        creator_column_name(["Год", "Средняя зарплата", "Средняя зарплата - Программист",
                                             "Количество вакансий", "Количество вакансий - Программист"], self.sheet_year)
        for year in self.dataSet_list[0].keys():
            self.sheet_year.append([year, self.dataSet_list[0][year], self.dataSet_list[1][year],
                               self.dataSet_list[2][year], self.dataSet_list[3][year]])
        setting_column_width_and_border_style(self.sheet_year)

        creator_column_name(["Город", "Уровень зарплат", " ", "Город", "Доля вакансий"], self.sheet_city)
        for index, key in enumerate(self.dataSet_list[4].keys()):
            self.sheet_city.append([key, self.dataSet_list[4][key], None, list(self.dataSet_list[5].keys())[index],
                               self.dataSet_list[5][list(self.dataSet_list[5].keys())[index]]])
        setting_column_width_and_border_style(self.sheet_city)

        for cell in self.sheet_city['E']:
            cell.number_format = FORMAT_PERCENTAGE_00
        self.workbook.save("report.xlsx")

    def generate_image(self):
        """
        Создает графики статистики в img

        :return: img
        """


        def creator_graph(number, value_1, value_2, name, bar_name_first, bar_name_second):
            """
            Создает один из видов графиков

            :param number: позиция в таблице
            :type number: int

            :param value_1: данные из vacancies
            :type value_1: dict

            :param value_2: данные из vacancies
            :type value_2:dict

            :param name: название графика
            :type name:str

            :param bar_name_first: название элемента графика
            :type bar_name_first:str

            :param bar_name_second: название элемента графика
            :type bar_name_second:str

            :return: png file
            """
            ax = fig.add_subplot(number)
            ax.set_title(name)
            ax.bar(x - 0.4 / 2, value_1.values(), 0.4, label=bar_name_first)
            ax.bar(x + 0.4 / 2, value_2.values(), 0.4, label=bar_name_second)
            ax.set_xticks(x, value_1.keys(), rotation="vertical")
            ax.legend(loc='best')
            ax.grid(True, axis='y')
            ax.legend(fontsize=8)

        fig = plt.figure()
        x = np.arange(len(self.dataSet_list[0].keys()))
        creator_graph(221, self.dataSet_list[0], self.dataSet_list[1],
                      "Уровень зарплат по годам", "средняя з/п", "з/п программист")
        creator_graph(222, self.dataSet_list[2], self.dataSet_list[3],
                      "Количество вакансии по годам", "Количество вакансии", "Количество вакансии\nпрограммист")

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        y_pos = np.arange(len(self.dataSet_list[4].keys()))
        ax.barh(y_pos, list(self.dataSet_list[4].values()), align='center')
        ax.set_yticks(y_pos, labels=self.dataSet_list[4].keys())
        ax.invert_yaxis()
        ax.grid(True, axis='x')

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансии по годам")
        keys = list(self.dataSet_list[5].keys())
        keys.append("Другие")
        values = list(self.dataSet_list[5].values())
        sum_values = sum(list(self.dataSet_list[5].values()))
        values.append(1 - sum_values)
        ax.pie(values, labels=keys, startangle=-5)

        fig.set_size_inches(9, 7)
        plt.tight_layout()
        plt.savefig("graph.png")

    def generate_pdf(self):
        """
        Создает pdf, где соединяется таблица статистики и графики статистики в img

        :return: pdf file
        """
        self.generate_excel()
        self.generate_image()
        for row in range(2, self.sheet_city.max_row + 1):
            for col in range(4, 6):
                if type(self.sheet_city.cell(row, col).value).__name__ == "float":
                    self.sheet_city.cell(row, col).value = str(round(self.sheet_city.cell(row, col).value * 100, 2)) + '%'

        pdf_template = Environment(loader=FileSystemLoader('.'))\
            .get_template("pdf_template.html").render({'name': 'Программист', 'png': "graph.png",
                                        'sheet_year':  self.sheet_year,
                                        'sheet_city': self.sheet_city})
        config = pdfkit.configuration(wkhtmltopdf=r'E:\загрузки\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={'enable-local-file-access': None})


start_time = datetime.now()
Report().generate_excel()
print(f"Total time: {datetime.now() - start_time}")