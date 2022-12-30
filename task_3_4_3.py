import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
import pandas as pd


class DataSet:
    """
    Класс для хранения списка вакансий.
    Attributes:
        file_name (str): Название файла
        vacancies_objects (list): Список вакансий
    """

    def __init__(self):
        """
        Конструктор для инициализация объекта DataSet, который создает поле для хранения списка вакансий
        Args:
             file_name (str): Название файла
        """
        self.salary_by_year = dict()
        self.vacancies_count_by_year = dict()
        self.salary_by_profession_name = dict()
        self.vacancies_count_by_profession_name = dict()
        self.salary_by_city = dict()
        self.vacancy_rate_by_city = dict()
        self.dict_lict = list()

    def return_list(self):
        return [self.salary_by_year, self.vacancies_count_by_year, self.salary_by_profession_name,
                self.vacancies_count_by_profession_name, self.salary_by_city, self.vacancy_rate_by_city]

class InputConnect:
    """Класс для ввода данных и формирования отчетности о вакансиях
    Args:
        params (tuple): Кортеж с названием файла и профессии
    """

    def __init__(self):
        """Конструктор для инициализации объекта InputConnect"""
        self.file_name, self.profession_name = InputConnect.get_params()

    @staticmethod
    def get_params():
        """Статический метод для ввода данные о вакансии
        :return: Кортеж с названием файла и профессии
        """
        file_name = input("Введите название файла: ")
        profession_name = input("Введите название профессии: ")
        return file_name, profession_name

    @staticmethod
    def print_data_dict(self, data: DataSet):
        """Вычисляет и печатает в консоль словари со статистикой о вакансиях
        :param self: Объект класса InputConnect
        :param data: Объект класса DataSet
        """
        df = pd.read_csv(self.file_name)
        count = len(df)
        df["salary"] = df[["salary_from", "salary_to"]].mean(axis=1)
        df["count"] = df.groupby("area_name")["area_name"].transform("count")
        df_norm = df[df["count"] > 0.01 * count]
        df_area = df_norm.groupby("area_name", as_index=False)["salary"].mean().sort_values(by="salary", ascending=False)
        df_area["salary"] = df_area["salary"].apply(lambda x: int(x))
        df_area10 = df_area.head(10)
        data.salary_by_city = dict(zip(df_area10["area_name"], df_area10["salary"]))

        data.vacancy_rate_by_city = {k: round(v / count, 4) for k, v in dict(df["area_name"].value_counts()).items()}

        print(f"Динамика уровня зарплат по годам: {data.salary_by_year}")
        print(f"Динамика количества вакансий по годам: {data.vacancies_count_by_year}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {data.salary_by_profession_name}")
        print(
            f"Динамика количества вакансий по годам для выбранной профессии: {data.vacancies_count_by_profession_name}")
        print(f"Уровень зарплат по городам (в порядке убывания): {data.salary_by_city}")
        print(f"Доля вакансий по городам (в порядке убывания): {dict(list(data.vacancy_rate_by_city.items())[:10])}")

    def read_csv_by_path(self, path: str):
        """Метод для многопоточной обработки csv файлов при помощи модуля pandas
        :param path: Путь до csv файла с вакансиями
        :param data: Объект класса DataSet
        """
        df = pd.read_csv(path)
        df["salary"] = df[["salary_from", "salary_to"]].mean(axis=1)
        df["published_at"] = df["published_at"].apply(lambda d: int(d[:4]))
        year = df["published_at"][0]
        df_vacancy = df["name"].str.contains(self.profession_name)

        filter_by_year = df["published_at"] == year
        salary_by_year = (year, int(df[filter_by_year]["salary"].mean()))
        vacancies_count_by_year = (year, len(df[filter_by_year]))
        salary_by_profession_name = (year, int(df[df_vacancy & filter_by_year]["salary"].mean()))
        vacancies_count_by_profession_name = (year, len(df[df_vacancy & filter_by_year]))
        return salary_by_year, vacancies_count_by_year, salary_by_profession_name, vacancies_count_by_profession_name



class Report:
    """
    Класс представляющий отчет
    :param dataSet_list: лист словарей с данными
    :type dataSet_list: list
    :param sheet_year: exl лист с таблицей статистики по годам
    :type sheet_year: worksheet
    :param sheet_city: exl лист с таблицами статистики по городам
    :type sheet_city: worksheet
    """
    def __init__(self):
        """
        Инициализирует объект класса Report
        """
        self.dataSet_list = DataSet().return_list()
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.sheet_year = self.workbook.create_sheet("Статистика по годам")
        self.sheet_city = self.workbook.create_sheet("Статистика по городам")

    def generate_excel(self):
        """
        Создает таблицу статистики в xlsx файл

        :return: возвращиет xlsx файл
        """
        def creator_column_name(array, sheet):
            """
            Создает в определенный sheet имена column

            :param array: массив столбцов таблицы
            :type array: list

            :param sheet: exl лист книги
            :type array: worksheet

            :return: обозвал ячейки шапки таблицы
            """
            for index, column_name in enumerate(array):
                sheet.cell(row=1, column=index + 1, value=column_name).font = Font(bold=True)

        def setting_column_width_and_border_style(sheet_year) -> None:
            """
             В exl листе создает ширину ячеек и их дизайн

            :param sheet_year: exl лист с таблицей статистики по годам
            :type sheet_year: worksheet

            :return:
            """
            for column_cells in sheet_year.columns:
                length = 0
                for cell in column_cells:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    if cell.value is not None:
                        length = max(len(str(cell.value)), length)
                sheet_year.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        creator_column_name(["Год", "Средняя зарплата", "Средняя зарплата - Программист",
                                             "Количество вакансий", "Количество вакансий - Программист"], self.sheet_year)
        for year in self.dataSet_list[0].keys():
            self.sheet_year.append([year, self.dataSet_list[0][year], self.dataSet_list[1][year],
                               self.dataSet_list[2][year], self.dataSet_list[3][year]])
        setting_column_width_and_border_style(self.sheet_year)

        creator_column_name(["Город", "Уровень зарплат", " ", "Город", "Доля вакансий"], self.sheet_city)
        for index, key in enumerate(self.dataSet_list[4].keys()):
            self.sheet_city.append([key, self.dataSet_list[4][key], None, list(self.dataSet_list[5].keys())[index],
                               self.dataSet_list[5][list(self.dataSet_list[5].keys())[index]]])
        setting_column_width_and_border_style(self.sheet_city)

        for cell in self.sheet_city['E']:
            cell.number_format = FORMAT_PERCENTAGE_00
        self.workbook.save("report.xlsx")

    def generate_image(self):
        """
        Создает графики статистики в img

        :return: img
        """


        def creator_graph(number, value_1, value_2, name, bar_name_first, bar_name_second):
            """
            Создает один из видов графиков

            :param number: позиция в таблице
            :type number: int

            :param value_1: данные из vacancies
            :type value_1: dict

            :param value_2: данные из vacancies
            :type value_2:dict

            :param name: название графика
            :type name:str

            :param bar_name_first: название элемента графика
            :type bar_name_first:str

            :param bar_name_second: название элемента графика
            :type bar_name_second:str

            :return: png file
            """
            ax = fig.add_subplot(number)
            ax.set_title(name)
            ax.bar(x - 0.4 / 2, value_1.values(), 0.4, label=bar_name_first)
            ax.bar(x + 0.4 / 2, value_2.values(), 0.4, label=bar_name_second)
            ax.set_xticks(x, value_1.keys(), rotation="vertical")
            ax.legend(loc='best')
            ax.grid(True, axis='y')
            ax.legend(fontsize=8)

        fig = plt.figure()
        x = np.arange(len(self.dataSet_list[0].keys()))
        creator_graph(221, self.dataSet_list[0], self.dataSet_list[1],
                      "Уровень зарплат по годам", "средняя з/п", "з/п программист")
        creator_graph(222, self.dataSet_list[2], self.dataSet_list[3],
                      "Количество вакансии по годам", "Количество вакансии", "Количество вакансии\nпрограммист")

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        y_pos = np.arange(len(self.dataSet_list[4].keys()))
        ax.barh(y_pos, list(self.dataSet_list[4].values()), align='center')
        ax.set_yticks(y_pos, labels=self.dataSet_list[4].keys())
        ax.invert_yaxis()
        ax.grid(True, axis='x')

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансии по годам")
        keys = list(self.dataSet_list[5].keys())
        keys.append("Другие")
        values = list(self.dataSet_list[5].values())
        sum_values = sum(list(self.dataSet_list[5].values()))
        values.append(1 - sum_values)
        ax.pie(values, labels=keys, startangle=-5)

        fig.set_size_inches(9, 7)
        plt.tight_layout()
        plt.savefig("graph.png")

    def generate_pdf(self):
        """
        Создает pdf, где соединяется таблица статистики и графики статистики в img

        :return: pdf file
        """
        self.generate_excel()
        self.generate_image()
        for row in range(2, self.sheet_city.max_row + 1):
            for col in range(4, 6):
                if type(self.sheet_city.cell(row, col).value).__name__ == "float":
                    self.sheet_city.cell(row, col).value = str(round(self.sheet_city.cell(row, col).value * 100, 2)) + '%'

        pdf_template = Environment(loader=FileSystemLoader('.'))\
            .get_template("pdf_template_3_4_3.html").render({'name': 'Программист', 'png': "graph.png",
                                        'sheet_year':  self.sheet_year})
        config = pdfkit.configuration(wkhtmltopdf=r'E:\загрузки\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report_3_4_2.pdf', configuration=config, options={'enable-local-file-access': None})


if __name__ == '__main__':
    inputparam = InputConnect()
    dataset = DataSet()
    Report().generate_pdf()