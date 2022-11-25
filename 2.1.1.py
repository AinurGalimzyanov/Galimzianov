import csv, re, math
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from collections import Counter


currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }


class Salary:
    def __init__(self, dictionary):
        self.salary_in_rur = currency_to_rub[dictionary["salary_currency"]] * \
                        (math.floor(float(dictionary["salary_to"])) + math.floor(float(dictionary["salary_from"])))/2


class Vacancy:
    def __init__(self, dictionary: dict):
        self.dic = dictionary
        self.salary = Salary(dictionary)
        self.dic["year"] = int(dictionary["published_at"][:4])
        self.is_needed = dictionary["is_needed"]


class InputConect:
    def __init__(self):
        self.file = input("Введите название файла: ")
        self.prof = input("Введите название профессии: ")
        with open(self.file, "r", encoding='utf-8-sig', newline='') as test:
            unpacker = iter(csv.reader(test))
            if next(unpacker, "none") == "none":
                print("Пустой файл")
                exit(0)
            if next(unpacker, "none") == "none":
                print("Нет данных")
                exit(0)

class DataSet:
    def __init__(self):
        self.input_values = InputConect()
        self.csv_reader()
        self.csv_filter()
        self.years = []
        for i in self.filt_vac:
            self.years.append(i.dic["year"])
        self.years.sort()
        conut_vacs = len(self.filt_vac)
        self.wages_year, self.count_year = self.get_key_and_count(self.filt_vac, "year", False)
        self.year_required_wages, self.year_required_count = self.get_key_and_count(self.required_vac, "year", False)
        self.wages_area, self.count_area = self.get_key_and_count(self.filt_vac, "area_name", True)
        self.wages_area = dict(list(sorted(self.wages_area.items(), key=lambda x: x[1], reverse=True))[:10])
        self.price_area = {x: round(y / conut_vacs, 4) for x, y in self.count_area.items()}
        self.price_area = dict(list(sorted(self.price_area.items(), key=lambda x: x[1], reverse=True))[:10])
        self.project_list = self.creat_list()
        self.write_data()

    def csv_reader(self):
        with open(self.input_values.file, "r", encoding='utf-8-sig', newline='') as test:
            csv_file = csv.reader(test)
            self.first = next(csv_file)
            self.lines = []
            lenght = 0
            for row in csv_file:
                if lenght < len(row):
                    lenght = len(row)
                if not "" in row and lenght == len(row):
                    self.lines.append(row)

    def csv_filter(self):
        self.filt_vac = []
        for i in self.lines:
            new_dict = dict(zip(self.first, i))
            new_dict["is_needed"] = new_dict["name"].find(self.input_values.prof) > -1
            self.filt_vac.append(Vacancy(new_dict))
        self.required_vac = list(filter(lambda x: x.is_needed, self.filt_vac))

    def try_add(self, dic: dict, key, x):
        try: dic[key] += x
        except: dic[key] = x
        return dic

    def upgrad_keys(self, count_key):
        for i in self.years:
            if i in count_key.keys():
                continue
            count_key[i] = 0
        return count_key

    def get_key_and_count(self, list_vacs: list, str_key: str, is_area: bool):
        sum_key, count_key = {}, {}
        for x in list_vacs:
            sum_key = self.try_add(sum_key, x.dic[str_key], x.salary.salary_in_rur)
            count_key = self.try_add(count_key, x.dic[str_key], 1)
        if not is_area:
            sum_key = self.upgrad_keys(sum_key)
            count_key = self.upgrad_keys(count_key)
        else:
            count_key = dict(filter(lambda y: y[1] / len(list_vacs) > 0.01, count_key.items()))
        salary_key = {}
        for key, x in count_key.items():
            salary_key[key] = 0 if x == 0 else math.floor(sum_key[key] / x)
        key_to_middle_salary = salary_key
        return key_to_middle_salary, count_key
    def write_data(self):
        print("Динамика уровня зарплат по годам:", self.wages_year)
        print("Динамика количества вакансий по годам:", self.count_year)
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.year_required_wages)
        print("Динамика количества вакансий по годам для выбранной профессии:", self.year_required_count)
        print("Уровень зарплат по городам (в порядке убывания):", self.wages_area)
        print("Доля вакансий по городам (в порядке убывания):", self.price_area)

    def creat_list(self):
        return [self.wages_year, self.year_required_wages, self.count_year,
                self.year_required_count, self.wages_area, self.price_area]

class Report:
    def __init__(self):
        self.dataSet_list = DataSet().project_list

    def generate_excel(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheet_year = workbook.create_sheet("Статистика по годам")
        sheet_city = workbook.create_sheet("Статистика по городам")

        def creator_column_name(array, sheet):
            for index, column_name in enumerate(array):
                sheet.cell(row=1, column=index + 1, value=column_name).font = Font(bold=True)

        def setting_column_width_and_border_style(sheet_year):
            for column_cells in sheet_year.columns:
                length = 0
                for cell in column_cells:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    if cell.value is not None:
                        length = max(len(str(cell.value)), length)
                sheet_year.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        creator_column_name(["Год", "Средняя зарплата", "Средняя зарплата - Программист",
                                             "Количество вакансий", "Количество вакансий - Программист"], sheet_year)
        for year in self.dataSet_list[0].keys():
            sheet_year.append([year, self.dataSet_list[0][year], self.dataSet_list[1][year],
                               self.dataSet_list[2][year], self.dataSet_list[3][year]])
        setting_column_width_and_border_style(sheet_year)

        creator_column_name(["Город", "Уровень зарплат", " ", "Город", "Доля вакансий"], sheet_city)
        for index, key in enumerate(self.dataSet_list[4].keys()):
            sheet_city.append([key, self.dataSet_list[4][key], None, list(self.dataSet_list[5].keys())[index],
                               self.dataSet_list[5][list(self.dataSet_list[5].keys())[index]]])
        setting_column_width_and_border_style(sheet_city)

        for cell in sheet_city['E']:
            cell.number_format = FORMAT_PERCENTAGE_00
        workbook.save("report.xlsx")


Report().generate_excel()
